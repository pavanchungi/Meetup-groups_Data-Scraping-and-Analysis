## Script for performing all the data manipulations, pre-processing, data-mapping between multiple data sources (csv files)

## and accordingly, having various analysis, insights into the data, visualisations, etc.



import numpy as np

import pandas as pd

import re

import warnings

import datetime

import matplotlib.pyplot as plt

import pylab

from statistics import mode

import geopy.distance

import copy

import csv

import openpyxl

#import xslxwriter

import os



## Data Loading

pd.options.display.max_rows=1000

warnings.filterwarnings('ignore')

file_path = '/home/manish/Data/meetup_groups_clean_28MAR2017.xlsx'

df_metadata = pd.read_excel(file_path, sheet_name='META DATA')

df_comicmeetupmaster = pd.read_excel(file_path, sheet_name='COMICMEETUP_MASTER')

df_latlong = pd.read_excel(file_path, sheet_name='lat_long', header=None, names=['CITY', 'STATE', 'LAT', 'LONG'])

df_founders = pd.read_excel(file_path, sheet_name='FOUNDERS')

df_members = pd.read_excel(file_path, sheet_name='MEMBERS')



## Data Manipulation

df_metadata.columns = df_metadata.columns.str.replace(' ', '_')

df_comicmeetupmaster.columns = df_comicmeetupmaster.columns.str.lstrip()

df_comicmeetupmaster.columns = df_comicmeetupmaster.columns.str.replace(' ', '_')

df_latlong.columns = df_latlong.columns.str.replace(' ', '_')

df_founders.columns = df_founders.columns.str.replace(' ', '_')

df_members.columns = df_members.columns.str.replace(' ', '_')



##df_metadata

df_metadata.apply(lambda x: sum(x.isnull()))



## Except the following, rest all group URLs are unique

df_metadata.loc[df_metadata.GROUP_URLs == df_metadata.GROUP_URLs.value_counts().index[0],]



## difference between group URLS and MEMBER URLs :

## the only major difference between these 2 columns is the additional '/members' in the URL

## so not much information contained in this column, so dropping these columns

df_metadata['GROUP_MEMBER_DIFF'] = [y.replace(x, '') if not y is np.nan else np.nan

                                    for x, y in zip(df_metadata.GROUP_URLs, df_metadata.MEMBER_URLs)]



## Extracting Groups from the URLs in df_metadata

df_metadata['Group'] = [re.findall(r'(?:.com/)(.+)/', x)[0] for x in df_metadata.GROUP_URLs]



## Extracting Country,State and City for the links from df_metadata

country = [np.nan] * len(df_metadata)

state = [np.nan] * len(df_metadata)

city = [np.nan] * len(df_metadata)

for i, x in enumerate(df_metadata.GEO_LOCATION_URLs):

    if x is not np.nan:

        temp = ''.join(re.findall(r'(?:/cities/)(.+)/', x)).split('/')

        if (len(temp) == 3):

            country[i] = temp[0]

            state[i] = temp[1]

            city[i] = temp[2]

        if (len(temp) == 2):

            country[i] = temp[0]

            state[i] = temp[1]

df_metadata['COUNTRY'] = country

df_metadata['STATE'] = state

df_metadata['CITY'] = city



## df_comicmeetupmaster

df_comicmeetupmaster.loc[502] = df_comicmeetupmaster.loc[502].combine_first(df_comicmeetupmaster.loc[422])

df_comicmeetupmaster.drop(422, axis=0, inplace=True)

df_comicmeetupmaster = df_comicmeetupmaster.reset_index(drop=True)



df_comicmeetupmaster.TOTAL_MEMBERS = df_comicmeetupmaster.TOTAL_MEMBERS.str.strip('\n')



df_comicmeetupmaster['INITIAL_CHAR'] = [x[0].upper() for x in df_comicmeetupmaster.GROUP]

df_comicmeetupmaster.loc[df_comicmeetupmaster.INITIAL_CHAR.str.isdigit(), 'INITIAL_CHAR'] = '0-9'



df_comicmeetupmaster.TOTAL_MEMBERS.replace('Otaku', np.nan, inplace=True)

df_comicmeetupmaster.TOTAL_MEMBERS = [np.nan if x == 'nan' else int(x)

                                      for x in df_comicmeetupmaster.TOTAL_MEMBERS.astype(str)]



df_comicmeetupmaster['DATETIME'] = pd.to_datetime(df_comicmeetupmaster.FOUNDED_DATE)

df_comicmeetupmaster['YEAR'] = [str(x).split()[0].split('-')[0] for x in df_comicmeetupmaster.DATETIME]

df_comicmeetupmaster['MONTH'] = [str(x).split()[0].split('-')[1] if not x is pd.NaT else pd.NaT

                                 for x in df_comicmeetupmaster.DATETIME]



## Filling null values of GEO_LOCATION_2 and correspondingly to CITY in df_latlong:

# df_comicmeetupmaster.groupby('GEO_LOCATION_1').GEO_LOCATION_2.nunique()

state_city = {'Arlington': 'TX', 'Brighton': 'UK', 'North Hampton': 'UK', 'Portsmouth': 'UK', 'Cape Cod': 'MA',

              'Fayetteville': 'NC', 'New Bury Park': 'CA', 'Niceville': 'FL', 'Springfield': 'IL', 'Stamford': 'CT',

              'Westchester': 'NY', 'Winchester': 'VA', 'Youngstown': 'OH'}



for i, row in df_comicmeetupmaster.iterrows():

    if ((row.GEO_LOCATION_2 is np.nan) & (row.GEO_LOCATION_1 in state_city.keys())):

        df_comicmeetupmaster.GEO_LOCATION_2[i] = state_city[row.GEO_LOCATION_1]



t = df_comicmeetupmaster.GEO_LOCATION_1 == 'Torrance, CA'

df_comicmeetupmaster.loc[t, 'GEO_LOCATION_2'] = df_comicmeetupmaster.loc[t, 'GEO_LOCATION_1'].values[0].split(',')[

    1].strip()

df_comicmeetupmaster.loc[t, 'GEO_LOCATION_1'] = df_comicmeetupmaster.loc[t, 'GEO_LOCATION_1'].values[0].split(',')[

    0].strip()

df_comicmeetupmaster.loc[df_comicmeetupmaster.GEO_LOCATION_2 == 'Manitoba', 'GEO_LOCATION_2'] = 'MB'



for i, row in df_latlong.iterrows():

    if ((row.STATE is np.nan) & (row.CITY in state_city.keys())):

        df_latlong.STATE[i] = state_city[row.CITY]



t = df_latlong.CITY == 'Torrance, CA'

df_latlong.loc[t, 'STATE'] = df_latlong.loc[t, 'CITY'].values[0].split(',')[1].strip()

df_latlong.loc[t, 'CITY'] = df_latlong.loc[t, 'CITY'].values[0].split(',')[0].strip()

df_latlong.loc[df_latlong.STATE == 'Manitoba', 'STATE'] = 'MB'



df_comicmeetupmaster.loc[

    (df_comicmeetupmaster.GEO_LOCATION_1 == 'Manchester') & (df_comicmeetupmaster.GEO_LOCATION_2.isnull()),

    'GEO_LOCATION_2'] = mode(df_latlong.loc[(df_latlong.CITY == 'Manchester'), 'STATE'].dropna())

df_comicmeetupmaster.loc[

    (df_comicmeetupmaster.GEO_LOCATION_1 == 'Kansas City') & (df_comicmeetupmaster.GEO_LOCATION_2.isnull()),

    'GEO_LOCATION_2'] = mode(df_latlong.loc[(df_latlong.CITY == 'Kansas City'), 'STATE'].dropna())

df_comicmeetupmaster.loc[df_comicmeetupmaster.GEO_LOCATION_1 == 'Vienna',

                         'GEO_LOCATION_2'] = df_latlong.loc[df_latlong.CITY == 'Vienna', 'STATE'].values[0]



## Fiiling up the missing Lat and Long in comicmeetupmaster



# for g in df_comicmeetupmaster.GEO_LOCATION_1.dropna().unique().tolist():

#    cond1 = (df_comicmeetupmaster.GEO_LOCATION_1==g)&(df_comicmeetupmaster.LATITUDE.isnull())

#    cond2 = (df_latlong.CITY==g)&(df_latlong.LAT.notnull())

#    if(any(cond1) and any(cond2)):

#        try:

#            df_comicmeetupmaster.loc[cond1,'LATITUDE'] = mode(df_latlong.loc[cond2,'LAT'].dropna())

#            df_comicmeetupmaster.loc[cond1,'LONGITUDE'] = mode(df_latlong.loc[cond2,'LONG'].dropna())

#        except:

#            pass

warnings.filterwarnings('ignore')

for i, row in df_comicmeetupmaster.iterrows():

    if ((row.GEO_LOCATION_1 is not np.nan) & (np.isnan(row.LATITUDE)) & (np.isnan(row.LONGITUDE))):

        df_comicmeetupmaster.LATITUDE[i] = mode(df_latlong.loc[(df_latlong.CITY == row.GEO_LOCATION_1) &

                                                               (

                                                               df_latlong.STATE == row.GEO_LOCATION_2), 'LAT'].dropna())

        df_comicmeetupmaster.LONGITUDE[i] = mode(df_latlong.loc[(df_latlong.CITY == row.GEO_LOCATION_1) &

                                                                (

                                                                df_latlong.STATE == row.GEO_LOCATION_2), 'LONG'].dropna())



## extracting the gender from the group name from df_comicmeetupmaster:

# female - Girl, Girls, Ladies, Her, Girlsonly, GeekGirlCon, Mommies, Queens, Women, Chic, Woman, Gals,

# male - Bros,

# others - Gay, Gays, PRIDE, LGBT, Lesbians, LesbianBi, LGBTQ, LGBTQIA, LGBTQA,

female = ['girl', 'ladies', 'mommies', 'queens', 'women', 'woman']

male = ['bros']

others = ['gay', 'pride', 'lgbt', 'lesbian']

female_mod = ['her', 'chic', 'gals']

female_mod = re.compile(r'\b(?:%s)\b' % '|'.join(female_mod))

df_comicmeetupmaster['GENDER'] = None

warnings.filterwarnings('ignore')



for i, x in enumerate(df_comicmeetupmaster.GROUP):

    if (any([temp in x.lower() for temp in female])):

        df_comicmeetupmaster.GENDER[i] = 'Female'

    elif (re.search(female_mod, x.lower())):

        df_comicmeetupmaster.GENDER[i] = 'Female'

    elif (any([temp in x.lower() for temp in male])):

        df_comicmeetupmaster.GENDER[i] = 'Male'

    elif (any([temp in x.lower() for temp in others])):

        df_comicmeetupmaster.GENDER[i] = 'Others'

    else:

        df_comicmeetupmaster.GENDER[i] = 'Unknown'



female_groups = pd.DataFrame(df_comicmeetupmaster.loc[df_comicmeetupmaster.GENDER == 'Female', 'GROUP']).reset_index(

    drop=True)

male_groups = pd.DataFrame(df_comicmeetupmaster.loc[df_comicmeetupmaster.GENDER == 'Male', 'GROUP']).reset_index(

    drop=True)

others_groups = pd.DataFrame(df_comicmeetupmaster.loc[df_comicmeetupmaster.GENDER == 'Others', 'GROUP']).reset_index(

    drop=True)

# Total 502 groups present



## df_founders

## Dropping those rows which have null values in both the Founded Date and Founder

for g in df_founders.Group.value_counts().index[(df_founders.Group.value_counts() > 1)]:

    for i in df_founders.loc[df_founders.Group == g,].index:

        if (sum(df_founders.loc[i, ['Founded_Date', 'Founder']].isnull()) == 2):

            df_founders.drop(i, axis=0, inplace=True)



df_founders.loc[527] = df_founders.loc[527].combine_first(df_founders.loc[447])

df_founders.drop(447, axis=0, inplace=True)

df_founders = df_founders.reset_index(drop=True)



## Cleaning the Date column

df_founders.Founded_Date = df_founders.Founded_Date.str.replace('Founded\n', '')

df_founders.Founded_Date = df_founders.Founded_Date.str.replace('Founded ', '')



### df_members is of no significance



### ANALYSIS RESULTS:



## Count of number of groups by state/country and city

q1a = df_comicmeetupmaster.groupby(by=['GEO_LOCATION_2', 'GEO_LOCATION_1'], as_index=False).agg({'GROUP': 'count'})

## Count of number of group by state/country only

q1b = df_comicmeetupmaster.groupby(by=['GEO_LOCATION_2'], as_index=False).agg({'GROUP': 'count'})





## Gender of the Group

q2 = pd.DataFrame(df_comicmeetupmaster.GENDER.value_counts()).reset_index()

q2.columns = ['GENDER', 'NUMBER']



## count of Groups for every alphabet

q3 = df_comicmeetupmaster.groupby('INITIAL_CHAR').agg({'GROUP': 'count'})





## Total members

## The total number of members are 407

len(df_members)



## Total groups

## The total number of groups are 502

len(df_founders.Group.unique())





## Count of average members in group by state/country and city

q6a = df_comicmeetupmaster.groupby(by=['GEO_LOCATION_2', 'GEO_LOCATION_1'], as_index=False).agg(

    {'TOTAL_MEMBERS': 'mean'})

q6a.rename(columns={'mean': 'AVERAGE MEMBERS'})

## Count of average number of group by state/country only

q6b = df_comicmeetupmaster.groupby(by=['GEO_LOCATION_2'], as_index=False).agg({'TOTAL_MEMBERS': 'mean'})

q6b.rename(columns={'mean': 'AVERAGE MEMBERS'})



## Total Lifecycle of this Market

## 14 years (from 15-Nov 2002 to 17-Aug 2016)

time_diff = max(df_comicmeetupmaster.DATETIME) - min(df_comicmeetupmaster.DATETIME)

q7 = round((time_diff.days / 30) / 12, 3)



## Average Startup duration between Groups

temp_date = df_comicmeetupmaster.DATETIME

temp_date = temp_date.sort_values()

temp_date.reset_index(drop=True, inplace=True)

temp_date.dropna(inplace=True)



sum_days = 0

avg_day = []

avg_day.append(0)

for i in range(len(temp_date)):

    if (i < (len(temp_date) - 1)):

        temp = (temp_date[i + 1] - temp_date[i])

        days = round((temp.days + (temp.seconds / (60 * 60 * 24))), 3)

        sum_days += days

        avg_day.append(days)

avg_open_distance = round(sum_days / len(temp_date), 3)



## From this even though the mean is 10 days, however this is very skewed data.

## This is due to the fact that many meetup groups that have been formed in 2016(149 groups)

## all are within less number of days from each other, as evident from graph

plt.plot(avg_day)

plt.xlabel('Difference between forming of two consecutive groups in days')

plt.ylabel('Number of Groups')

plt.show()



plt.plot(temp_date, avg_day)

plt.xlabel('Date of forming of the meetup groups')

plt.ylabel('Average number of days')

plt.show()



## From this even the median is 3 days and 75th percentile is 7

pd.Series(avg_day).describe()



## Yearly Average Startup duration between groups

yr_unique = df_comicmeetupmaster.YEAR.unique().tolist()

yr_unique.remove('NaT')

yr_unique = sorted(yr_unique)



final = []

for y in yr_unique:

    temp_date = df_comicmeetupmaster.loc[df_comicmeetupmaster.YEAR == y, 'DATETIME']

    temp_date = temp_date.sort_values()

    temp_date.reset_index(drop=True, inplace=True)

    temp_date.dropna(inplace=True)



    sum_days = 0

    for i in range(len(temp_date)):

        if (i < (len(temp_date) - 1)):

            temp = (temp_date[i + 1] - temp_date[i])

            sum_days += round((temp.days + (temp.seconds / (60 * 60 * 24))), 3)

    avg_open_distance = round(sum_days / len(temp_date), 3)

    final.append([y, avg_open_distance])



q8 = pd.DataFrame(final, columns=['Year', 'Average Opening'])



## For the groups in the same city, lat and long co-ords are exactly the same, so doing it on country wise

final = []

for country in df_comicmeetupmaster.GEO_LOCATION_2.dropna().unique().tolist():

    len_cities = len(df_comicmeetupmaster.loc[df_comicmeetupmaster.GEO_LOCATION_2 == country, 'GEO_LOCATION_1'])

    temp_cities = df_comicmeetupmaster.loc[

        df_comicmeetupmaster.GEO_LOCATION_2 == country, 'GEO_LOCATION_1'].dropna().unique().tolist()

    if len(temp_cities) > 1:

        l = []

        for city in temp_cities:

            temp_lat = df_comicmeetupmaster.loc[(df_comicmeetupmaster.GEO_LOCATION_2 == country) &

                                                (df_comicmeetupmaster.GEO_LOCATION_1 == city),].LATITUDE.unique()

            temp_long = df_comicmeetupmaster.loc[(df_comicmeetupmaster.GEO_LOCATION_2 == country) &

                                                 (df_comicmeetupmaster.GEO_LOCATION_1 == city),].LONGITUDE.unique()

            l.append([temp_lat[0], temp_long[0]])



        sum_temp = 0



        for i in range(len(l) - 1):

            tp = l[i]

            new_l = l[-(len(l) - 1 - i):]



            for j in range(len(new_l)):

                temp = round(geopy.distance.distance(tp, new_l[j]).km, 3)

                sum_temp += temp

        avg_distance = round(sum_temp / len_cities, 3)

        final.append([country, avg_distance, len_cities, len(temp_cities)])



q9 = pd.DataFrame(final, columns=['STATE', 'AVG_DISTANCE', 'TOTAL_GROUPS', 'TOTAL_CITIES'])



# Group categories/genres

genre_dict = {'Comics': ['comic', 'cartoon', 'anime', 'manga', 'genshiken', 'brony'], 'Geeks': ['geek', 'nerd'],

              'Action': ['action', 'adventure', 'game', 'gaming', 'juego', 'pokemon'], 'Graphic': ['graphic'],

              'Fiction': ['scifi', 'fiction', 'star war', 'star trek', 'superhero', 'sci fi', 'science fiction'],

              'Doctor': ['doctor'], 'Vintage': ['back to the 80s'], 'Pop': ['pop culture'], 'Black': ['black'],

              'Movies': ['movie', 'film'], 'Trade': ['trade', 'sell', 'guild', 'crafter'],

              'Arts': ['drawing', 'writing', 'dance', 'artist', 'music', 'sword', 'arts', 'sketch', 'gym', 'culture',

                       'design', 'cosplay']}



final_dict = {}

for k in genre_dict.keys():

    final_dict[k] = []

    for x in df_comicmeetupmaster.GROUP:

        if (any([temp in x.lower() for temp in genre_dict[k]])):

            final_dict[k].append(x)



arts_mod = re.compile(r'\b(?:%s)\b' % '|'.join(['art']))

for x in df_comicmeetupmaster.GROUP:

    if (re.search(arts_mod, x.lower())):

        final_dict['Arts'].append(x)



if os.path.exists('Answers.xlsx'):

    os.remove('Answers.xlsx')



# book = openpyxl.load_workbook('Answers.xlsx')

book = openpyxl.Workbook()

defaultsheet = book['Sheet']

book.remove(defaultsheet)



for s in ['1a', '1b', 2, 3, '6a', '6b', 8, 9]:

    book.create_sheet('Question {0}'.format(s))



book.save('Answers.xlsx')

# book.sheetnames



writer = pd.ExcelWriter('Answers.xlsx', engine='xlsxwriter')



q1a.to_excel(writer, sheet_name='Question 1a')

q1b.to_excel(writer, sheet_name='Question 1b')

q2.to_excel(writer, sheet_name='Question 2')

q3.to_excel(writer, sheet_name='Question 3')

q6a.to_excel(writer, sheet_name='Question 6a')

q6b.to_excel(writer, sheet_name='Question 6b')

q8.to_excel(writer, sheet_name='Question 8')

q9.to_excel(writer, sheet_name='Question 9')



writer.save()



temp = pd.DataFrame(

    {'GENRE': [k for k in final_dict.keys()], 'GROUP COUNT': [len(final_dict[k]) for k in final_dict.keys()]})

temp.index += 1

temp.to_csv('Question10.csv')



with open('Question10.csv', 'a') as f:

    f.write('\n\n\n')

    for key in final_dict.keys():

        f.write('{0}\n'.format(key))

        for val in final_dict[key]:

            f.write('{0}\n'.format(val))

        f.write('\n')
